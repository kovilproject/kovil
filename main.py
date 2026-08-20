import base64
from datetime import date, datetime
import io
import os
import sqlite3
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
import pandas as pd
from reportlab.lib.pagesizes import A4
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from reportlab.pdfgen import canvas
import streamlit as st
import streamlit.components.v1 as components

# ---------------------------------------------------------
# STREAMLIT PAGE CONFIG & SESSION INITIALIZATION
# ---------------------------------------------------------
st.set_page_config(
    page_title="அருள்மிகு பெத்தையா காடேரி அம்பிகை", page_icon="🛕", layout="wide"
)

query_params = st.query_params

if "logged_in" not in st.session_state:
    if query_params.get("logged_in") == "true":
        st.session_state.logged_in = True
        st.session_state.username = query_params.get("user", "admin")
    else:
        st.session_state.logged_in = False
        st.session_state.username = ""

# Mobile Viewport & Thermal Print Specific CSS
st.markdown(
    """
    <style>
    body, .stApp {
        overscroll-behavior-y: contain;
    }

    /* Thermal Print Styling */
    @media print {
        body * {
            visibility: hidden;
        }
        #thermal-receipt, #thermal-receipt * {
            visibility: visible;
        }
        #thermal-receipt {
            position: absolute;
            left: 0;
            top: 0;
            width: 80mm; /* 80mm Thermal Printer Standard */
            font-family: monospace, sans-serif;
            font-size: 12px;
            padding: 5mm;
            color: black !important;
        }
    }
    </style>
    """,
    unsafe_allow_html=True,
)


# ---------------------------------------------------------
# BACKGROUND SETUP
# ---------------------------------------------------------
def set_local_background(image_path):
    with open(image_path, "rb") as image_file:
        encoded_string = base64.b64encode(image_file.read()).decode()

    bg_css = f"""
    <style>
    .stApp {{
        background-image: linear-gradient(rgba(255, 255, 255, 0.88), rgba(255, 255, 255, 0.88)), url("data:image/png;base64,{encoded_string}");
        background-size: cover;
        background-position: center;
        background-attachment: fixed;
    }}
    </style>
    """
    st.markdown(bg_css, unsafe_allow_html=True)


if os.path.exists("bg.jpg"):
    set_local_background("bg.jpg")

# ---------------------------------------------------------
# DATABASE SETUP (SQLite)
# ---------------------------------------------------------
DB_NAME = "kovil_kanakku.db"


def get_db_connection():
    return sqlite3.connect(DB_NAME)


def init_db():
    with get_db_connection() as conn:
        cursor = conn.cursor()
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS receipts (
                receipt_no INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                category TEXT,
                name TEXT,
                city TEXT,
                amount REAL,
                phone TEXT,
                payment_method TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS expenses (
                expense_id INTEGER PRIMARY KEY AUTOINCREMENT,
                date TEXT,
                category TEXT,
                title TEXT,
                amount REAL,
                remarks TEXT
            )
        """)
        cursor.execute("""
            CREATE TABLE IF NOT EXISTS users (
                username TEXT PRIMARY KEY,
                password TEXT
            )
        """)

        cursor.execute("PRAGMA table_info(receipts)")
        columns = [column[1] for column in cursor.fetchall()]
        if "phone" not in columns:
            cursor.execute("ALTER TABLE receipts ADD COLUMN phone TEXT")
        if "payment_method" not in columns:
            cursor.execute(
                "ALTER TABLE receipts ADD COLUMN payment_method TEXT DEFAULT"
                " 'Cash (பணம்)'"
            )

        cursor.execute("SELECT * FROM users WHERE username = 'admin'")
        if not cursor.fetchone():
            cursor.execute(
                "INSERT INTO users (username, password) VALUES (?, ?)",
                ("admin", "kovil123"),
            )
        conn.commit()


init_db()


# ---------------------------------------------------------
# HELPER FUNCTIONS (PDF, EXCEL & THERMAL HTML GENERATION)
# ---------------------------------------------------------
def render_thermal_receipt_html(data):
    # data: (receipt_no, date, category, name, city, amount, phone, payment_method)
    r_no, r_date, r_cat, r_name, r_city, r_amt, r_phone, r_pay = data

    html_code = f"""
    <div id="thermal-receipt" style="width: 280px; font-family: 'Courier New', monospace; border: 1px dashed #000; padding: 10px; margin: auto; background: #fff; color: #000;">
        <div style="text-align: center; font-weight: bold; font-size: 14px;">
            அருள்மிகு பெத்தையா காடேரி அம்பிகை
        </div>
        <div style="text-align: center; font-size: 10px; margin-bottom: 5px;">
            மஞ்சள் நீராட்டு வெள்ளாள சமூக குலதெய்வ மண்டகப்படி
        </div>
        <hr style="border-top: 1px dashed #000;">
        <table style="width: 100%; font-size: 11px; text-align: left;">
            <tr><td><b>ரசீது எண்:</b> {r_no}</td><td style="text-align:right;"><b>தேதி:</b> {r_date}</td></tr>
        </table>
        <hr style="border-top: 1px dashed #000;">
        <div style="font-size: 11px; line-height: 1.5;">
            <b>வரவு வகை:</b> {r_cat}<br>
            <b>பெயர்:</b> {r_name}<br>
            <b>ஊர்:</b> {r_city if r_city else '-'}<br>
            <b>கைபேசி:</b> {r_phone if r_phone else '-'}<br>
            <b>செலுத்திய முறை:</b> {r_pay if r_pay else 'Cash'}<br>
        </div>
        <hr style="border-top: 1px dashed #000;">
        <div style="text-align: center; font-size: 15px; font-weight: bold; margin: 5px 0;">
            தொகை: Rs. {r_amt:,.2f}/-
        </div>
        <hr style="border-top: 1px dashed #000;">
        <div style="text-align: center; font-size: 10px; margin-top: 5px;">
            நன்றி! அருள்மிகு பெத்தையா காடேரி அம்பிகை துணை!
        </div>
    </div>
    <br>
    <div style="text-align: center;">
        <button onclick="window.print()" style="background-color: #04AA6D; color: white; padding: 8px 16px; border: none; border-radius: 4px; cursor: pointer; font-weight: bold;">
            🖨️ பிரிண்ட் செய்க (Print Now)
        </button>
    </div>
    """
    return html_code


def get_pdf_font():
    local_font = "NotoSansTamil-Regular.ttf"
    if os.path.exists(local_font):
        try:
            pdfmetrics.registerFont(TTFont("TamilFont", local_font))
            return "TamilFont"
        except Exception:
            pass

    system_fonts = [
        "C:\\Windows\\Fonts\\Nirmala.ttf",
        "C:\\Windows\\Fonts\\latha.ttf",
    ]
    for font_path in system_fonts:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont("TamilFont", font_path))
                return "TamilFont"
            except Exception:
                continue
    return "Helvetica"


def generate_receipt_pdf(
        title, name, city, amount, receipt_no, date_str, phone, pay_method
):
    buffer = io.BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4
    font_name = get_pdf_font()

    c.setLineWidth(2)
    c.rect(20, 20, width - 40, height - 40)

    c.setFont(font_name, 20)
    c.drawCentredString(
        width / 2, height - 80, "அருள்மிகு பெத்தையா காடேரி அம்பிகை"
    )
    c.setFont(font_name, 12)
    c.drawCentredString(
        width / 2, height - 105, "மஞ்சள் நீராட்டு வெள்ளாள சமூக குலதெய்வ மண்டகப்படி"
    )

    c.setFont(font_name, 12)
    c.drawString(50, height - 175, f"ரசீது எண் :  {receipt_no}")
    c.drawString(width - 180, height - 175, f"தேதி :  {date_str}")

    c.setFont(font_name, 16)
    c.drawCentredString(width / 2, height - 210, f"வரவு வகை: {title}")

    c.rect(40, height - 460, width - 80, 230)
    c.setFont(font_name, 13)
    c.drawString(60, height - 260, f"பெயர் (Name)           :   {name}")
    c.drawString(
        60, height - 300, f"கைபேசி எண் (Phone)     :   {phone if phone else 'N/A'}"
    )
    c.drawString(60, height - 340, f"ஊர் / பகுதி (City)         :   {city}")
    c.drawString(
        60,
        height - 380,
        f"செலுத்திய முறை (Mode)  :   {pay_method if pay_method else 'Cash'}",
    )
    c.drawString(
        60, height - 420, f"தொகை (Amount)           :   Rs. {amount:,.2f}/-"
    )

    c.showPage()
    c.save()
    buffer.seek(0)
    return buffer


def generate_excel_report(rows, report_type, category, total_amt):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "அறிக்கை"

    title_font = Font(name="Calibri", size=16, bold=True, color="800000")
    sub_font = Font(name="Calibri", size=11, italic=True)
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    bold_font = Font(name="Calibri", size=11, bold=True)

    header_fill = PatternFill(
        start_color="4A0E17", end_color="4A0E17", fill_type="solid"
    )
    total_fill = PatternFill(
        start_color="EAEAEA", end_color="EAEAEA", fill_type="solid"
    )
    thin_border = Border(
        left=Side(style="thin", color="D3D3D3"),
        right=Side(style="thin", color="D3D3D3"),
        top=Side(style="thin", color="D3D3D3"),
        bottom=Side(style="thin", color="D3D3D3"),
    )

    ws.merge_cells("A1:H1")
    ws["A1"] = "அருள்மிகு பெத்தையா காடேரி அம்பிகை"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:H2")
    ws["A2"] = "மஞ்சள் நீராட்டு வெள்ளாள சமூக குலதெய்வ மண்டகப்படி"
    ws["A2"].font = sub_font
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A3:H3")
    ws["A3"] = (
        f"அறிக்கை வகை: {report_type} | பிரிவு: {category} | உருவாக்கப்பட்ட"
        f" தேதி: {datetime.now().strftime('%d-%m-%Y')}"
    )
    ws["A3"].font = bold_font
    ws["A3"].alignment = Alignment(horizontal="center")

    ws.append([])
    is_income = "வரவு" in report_type
    id_header = "ரசீது எண்" if is_income else "செலவு எண்"

    if is_income:
        headers = [
            "வ.எண் (S.No)",
            id_header,
            "தேதி",
            "வகை",
            "கைபேசி",
            "பெயர் / விவரம்",
            "செலுத்திய முறை",
            "தொகை (₹)",
        ]
    else:
        headers = [
            "வ.எண் (S.No)",
            id_header,
            "தேதி",
            "வகை",
            "குறிப்பு",
            "விவரம்",
            "-",
            "தொகை (₹)",
        ]

    ws.append(headers)

    header_row = 5
    for col_num in range(1, 9):
        cell = ws.cell(row=header_row, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    current_row = 6
    for idx, row in enumerate(rows, start=1):
        p_val = row[5] if len(row) > 5 and row[5] else "-"
        p_method = (
            row[6]
            if len(row) > 6 and row[6]
            else ("-" if not is_income else "Cash (பணம்)")
        )

        ws.append(
            [idx, row[0], row[1], row[2], p_val, row[3], p_method, row[4]]
        )
        for col_num in range(1, 9):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if col_num == 8:
                cell.number_format = "₹#,##0.00"
                cell.alignment = Alignment(horizontal="right")
            elif col_num in [1, 2, 3, 5, 7]:
                cell.alignment = Alignment(horizontal="center")
        current_row += 1

    ws.merge_cells(
        start_row=current_row,
        start_column=1,
        end_row=current_row,
        end_column=7,
    )
    ws.cell(row=current_row, column=1, value="மொத்தம்").font = bold_font
    ws.cell(row=current_row, column=1).alignment = Alignment(
        horizontal="right"
    )

    tot_cell = ws.cell(row=current_row, column=8, value=total_amt)
    tot_cell.font = bold_font
    tot_cell.number_format = "₹#,##0.00"
    tot_cell.alignment = Alignment(horizontal="right")

    for col_num in range(1, 9):
        c = ws.cell(row=current_row, column=col_num)
        c.fill = total_fill
        c.border = thin_border

    column_widths = {
        "A": 12,
        "B": 14,
        "C": 15,
        "D": 22,
        "E": 18,
        "F": 30,
        "G": 20,
        "H": 20,
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ---------------------------------------------------------
# AUTHENTICATION (LOGIN & SIGN UP)
# ---------------------------------------------------------
def login():
    st.title("🛕 அருள்மிகு பெத்தையா காடேரி அம்பிகை")

    login_tab, signup_tab = st.tabs(
        ["🔓 உள்நுழைக (Login)", "📝 புதிய கணக்கு உருவாக்க (Sign Up)"]
    )

    with login_tab:
        st.subheader("கணக்கு மேலாண்மை - உள்நுழைவு")
        with st.form("login_form"):
            username = st.text_input("பயனர் பெயர் (Username)")
            password = st.text_input("கடவுச்சொல் (Password)", type="password")
            submit = st.form_submit_button(
                "🔓 உள்நுழைக (Login)", type="primary"
            )

            if submit:
                if not username or not password:
                    st.warning("⚠️ பயனர் பெயர் மற்றும் கடவுச்சொல்லை உள்ளிடவும்!")
                else:
                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT * FROM users WHERE username = ? AND"
                            " password = ?",
                            (username, password),
                        )
                        user = cursor.fetchone()

                    if user:
                        st.session_state.logged_in = True
                        st.session_state.username = username
                        st.query_params["logged_in"] = "true"
                        st.query_params["user"] = username
                        st.rerun()
                    else:
                        st.error("❌ தவறான பயனர் பெயர் அல்லது கடவுச்சொல்!")

    with signup_tab:
        st.subheader("புதிய பயனர் கணக்கு உருவாக்க")
        with st.form("signup_form", clear_on_submit=True):
            new_username = st.text_input("புதிய பயனர் பெயர் (New Username)")
            new_password = st.text_input(
                "புதிய கடவுச்சொல் (New Password)", type="password"
            )
            confirm_password = st.text_input(
                "கடவுச்சொல்லை உறுதிசெய்க (Confirm Password)", type="password"
            )
            signup_submit = st.form_submit_button(
                "📝 கணக்கு உருவாக்கு (Create Account)", type="primary"
            )

            if signup_submit:
                if not new_username or not new_password:
                    st.warning("⚠️ அனைத்து விவரங்களையும் நிரப்பவும்!")
                elif new_password != confirm_password:
                    st.error("❌ கடவுச்சொற்கள் பொருந்தவில்லை!")
                else:
                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "SELECT * FROM users WHERE username = ?",
                            (new_username,),
                        )
                        existing_user = cursor.fetchone()

                        if existing_user:
                            st.error("❌ இந்த பயனர் பெயர் ஏற்கனவே உள்ளது!")
                        else:
                            cursor.execute(
                                "INSERT INTO users (username, password) VALUES"
                                " (?, ?)",
                                (new_username, new_password),
                            )
                            conn.commit()
                            st.success(
                                "✅ புதிய கணக்கு உருவாக்கப்பட்டது! Login Tab-ல்"
                                " உள்நுழையலாம்."
                            )


# ---------------------------------------------------------
# MAIN APPLICATION INTERFACE
# ---------------------------------------------------------
if not st.session_state.logged_in:
    login()
else:
    st.sidebar.title(f"👤 வரவேற்பு: {st.session_state.username}")
    if st.sidebar.button("🚪 வெளியேறு (Logout)"):
        st.session_state.logged_in = False
        st.session_state.username = ""
        st.query_params.clear()
        st.rerun()

    st.title(
        "🛕 அருள்மிகு பெத்தையா காடேரி அம்பிகை மஞ்சள் நீராட்டு வெள்ளாள சமூக குலதெய்வ"
        " மண்டகப்படி"
    )

    tab1, tab2, tab3, tab4 = st.tabs(
        [
            "📥 புதிய வரவு (Receipt)",
            "📤 புதிய செலவு (Expense)",
            "📊 அறிக்கைகள் (Reports)",
            "✏️ பதிவைத் திருத்த (Edit Entry)",
        ]
    )

    # TAB 1: RECEIPT
    with tab1:
        st.header("📥 புதிய வரவு பதிவு")

        with st.form("receipt_form", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                receipt_date = st.date_input(
                    "தேதி (Date)",
                    value=date.today(),
                    min_value=date(1900, 1, 1),
                    max_value=date(2100, 12, 31),
                )
                category = st.selectbox(
                    "வரவு வகை",
                    [
                        "நிரந்தர வைப்பு நிதியாளர்கள் வரவுகள்",
                        "காணிக்கையாளர்கள் வரவுகள்",
                        "சிறப்பு வைப்பு நிதியாளர்கள் வரவுகள்",
                        "சிறப்பு காணிக்கையாளர்கள் வரவுகள்",
                        "திருவிளக்கு பூஜை வரவுகள்",
                    ],
                )
                phone = st.text_input("கைபேசி எண் (Phone No)")
                payment_method = st.selectbox(
                    "பணம் செலுத்திய முறை (Payment Method)",
                    [
                        "Cash (பணம்)",
                        "GPay / PhonePe (UPI)",
                        "Bank Transfer (வங்கி மாற்றம்)",
                         "Money order (மணி ஆர்டர் )",
                    ],
                )

            with col2:
                name = st.text_input("பெயர் (Name)")
                city = st.text_input("ஊர் / பகுதி (City)")
                amount = st.number_input(
                    "தொகை (₹)", min_value=0.0, step=100.0, format="%.2f"
                )

            submit_rec = st.form_submit_button(
                "💾 சேமி (Save Receipt)", type="primary"
            )

            if submit_rec:
                if not name or amount <= 0:
                    st.warning(
                        "⚠️ தயவுசெய்து பெயர் மற்றும் சரியான தொகையை நிரப்பவும்!"
                    )
                else:
                    formatted_date = receipt_date.strftime("%d-%m-%Y")

                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "INSERT INTO receipts (date, category, name, city,"
                            " amount, phone, payment_method) VALUES (?, ?,"
                            " ?, ?, ?, ?, ?)",
                            (
                                formatted_date,
                                category,
                                name,
                                city,
                                amount,
                                phone,
                                payment_method,
                            ),
                        )
                        conn.commit()
                        rec_id = cursor.lastrowid

                    st.session_state["last_receipt_id"] = rec_id
                    st.success(
                        "✅ வரவு வெற்றிகரமாகச் சேமிக்கப்பட்டது! (ரசீது எண்:"
                        f" {rec_id})"
                    )

        st.divider()
        st.subheader("🖨️ ரசீது அச்சிடுதல் (Thermal & A4 PDF)")

        default_r_no = st.session_state.get("last_receipt_id", 1)
        r_no_input = st.number_input(
            "ரசீது எண் உள்ளிடவும்:",
            min_value=1,
            step=1,
            value=int(default_r_no),
        )

        if st.button("🔍 ரசீது தேடு"):
            with get_db_connection() as conn:
                cursor = conn.cursor()
                cursor.execute(
                    "SELECT receipt_no, date, category, name, city, amount,"
                    " phone, payment_method FROM receipts WHERE receipt_no = ?",
                    (r_no_input,),
                )
                st.session_state["active_receipt"] = cursor.fetchone()

        if (
                "active_receipt" in st.session_state
                and st.session_state["active_receipt"]
        ):
            data = st.session_state["active_receipt"]

            col_p1, col_p2 = st.columns(2)

            with col_p1:
                st.markdown("### 📄 A4 Standard PDF")
                pdf_bytes = generate_receipt_pdf(
                    data[2],
                    data[3],
                    data[4],
                    data[5],
                    data[0],
                    data[1],
                    data[6],
                    data[7],
                )
                st.download_button(
                    label="📥 A4 PDF பதிவிறக்கு",
                    data=pdf_bytes,
                    file_name=f"Receipt_{data[0]}.pdf",
                    mime="application/pdf",
                )

            with col_p2:
                st.markdown("### 🧾 Thermal POS Receipt")
                thermal_html = render_thermal_receipt_html(data)
                components.html(thermal_html, height=320, scrolling=True)

    # TAB 2: EXPENSE
    with tab2:
        st.header("📤 புதிய செலவு பதிவு")

        with st.form("expense_form", clear_on_submit=True):
            col1, col2 = st.columns(2)
            with col1:
                expense_date = st.date_input(
                    "செலவு தேதி (Date)",
                    value=date.today(),
                    min_value=date(1900, 1, 1),
                    max_value=date(2100, 12, 31),
                )
                exp_category = st.selectbox(
                    "செலவு வகை",
                    [
                        "மின்சாரக் கட்டணம் (EB Bill)",
                        "பூஜைப் பொருட்கள்",
                        "அன்னதானம்",
                        "வேலை ஆள் கூலி",
                        "பராமரிப்புச் செலவு",
                        "இதர செலவுகள்",
                    ],
                )
                exp_title = st.text_input("விவரம் (Title)")
            with col2:
                exp_amount = st.number_input(
                    "செலவுத் தொகை (₹)",
                    min_value=0.0,
                    step=50.0,
                    format="%.2f",
                )
                exp_remarks = st.text_input("குறிப்பு (Remarks)")

            submit_exp = st.form_submit_button(
                "💾 செலவைச் சேமி (Save Expense)", type="primary"
            )

            if submit_exp:
                if not exp_title or exp_amount <= 0:
                    st.warning("⚠️ விவரம் மற்றும் சரியான தொகையை நிரப்பவும்!")
                else:
                    formatted_exp_date = expense_date.strftime("%d-%m-%Y")
                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        cursor.execute(
                            "INSERT INTO expenses (date, category, title,"
                            " amount, remarks) VALUES (?, ?, ?, ?, ?)",
                            (
                                formatted_exp_date,
                                exp_category,
                                exp_title,
                                exp_amount,
                                exp_remarks,
                            ),
                        )
                        conn.commit()
                    st.success("✅ செலவு பதிவு வெற்றிகரமாக சேமிக்கப்பட்டது!")

    # TAB 3: REPORTS
    with tab3:
        st.header("📊 கணக்கு அறிக்கைகள்")

        view_type = st.radio(
            "அறிக்கை வகை:",
            ["📥 வரவு பட்டியல் (Income)", "📤 செலவு பட்டியல் (Expense)"],
            horizontal=True,
        )
        cat_filter = st.selectbox(
            "வகை வடிகட்டி (Filter):",
            [
                "அனைத்தும் (All)",
                "நிரந்தர வைப்பு நிதியாளர்கள் வரவுகள்",
                "காணிக்கையாளர்கள் வரவுகள்",
                "சிறப்பு வைப்பு நிதியாளர்கள் வரவுகள்",
                "சிறப்பு காணிக்கையாளர்கள் வரவுகள்",
                "திருவிளக்கு பூஜை வரவுகள்",
                "மின்சாரக் கட்டணம் (EB Bill)",
                "பூஜைப் பொருட்கள்",
                "அன்னதானம்",
                "வேலை ஆள் கூலி",
                "பராமரிப்புச் செலவு",
                "இதர செலவுகள்",
            ],
        )

        with get_db_connection() as conn:
            cursor = conn.cursor()

            params = []
            if "வரவு" in view_type:
                query = (
                    "SELECT receipt_no, date, category, name, amount, phone,"
                    " payment_method FROM receipts"
                )
                if cat_filter != "அனைத்தும் (All)":
                    query += " WHERE category = ?"
                    params.append(cat_filter)
                query += " ORDER BY receipt_no ASC"
            else:
                query = (
                    "SELECT expense_id, date, category, title, amount, remarks"
                    " FROM expenses"
                )
                if cat_filter != "அனைத்தும் (All)":
                    query += " WHERE category = ?"
                    params.append(cat_filter)
                query += " ORDER BY expense_id ASC"

            cursor.execute(query, params)
            rows = cursor.fetchall()

            cursor.execute("SELECT SUM(amount) FROM receipts")
            tot_v = cursor.fetchone()[0] or 0.0
            cursor.execute("SELECT SUM(amount) FROM expenses")
            tot_s = cursor.fetchone()[0] or 0.0

        filtered_tot = sum(r[4] for r in rows) if rows else 0.0

        col_m1, col_m2 = st.columns(2)
        with col_m1:
            st.metric(
                label="தேர்ந்தெடுக்கப்பட்ட வகைத் தொகை",
                value=f"₹{filtered_tot:,.2f}",
            )
        with col_m2:
            st.metric(
                label="மொத்தக் கையிருப்பு (Balance)",
                value=f"₹{(tot_v - tot_s):,.2f}",
            )

        st.divider()

        # DELETE SECTION
        with st.expander("🗑️ பதிவை நீக்க (Delete Entry)"):
            st.warning(
                "⚠️ கவனிக்க: நீக்கப்பட்ட பதிவு டேட்டாபேஸிலிருந்து நிரந்தரமாக"
                " அழிக்கப்படும்!"
            )
            delete_id = st.number_input(
                "நீக்க வேண்டிய எண் (ID / Receipt No) உள்ளிடவும்:",
                min_value=1,
                step=1,
            )

            if st.button("❌ பதிவை நீக்கு (Confirm Delete)", type="primary"):
                with get_db_connection() as conn:
                    cursor = conn.cursor()
                    if "வரவு" in view_type:
                        cursor.execute(
                            "DELETE FROM receipts WHERE receipt_no = ?",
                            (delete_id,),
                        )
                    else:
                        cursor.execute(
                            "DELETE FROM expenses WHERE expense_id = ?",
                            (delete_id,),
                        )
                    deleted_count = cursor.rowcount
                    conn.commit()

                if deleted_count > 0:
                    st.success(
                        f"✅ எண் {delete_id} பதிவு வெற்றிகரமாக நீக்கப்பட்டது!"
                    )
                    st.rerun()
                else:
                    st.error(f"❌ எண் {delete_id} காணப்படவில்லை!")

        st.divider()

        # DATA TABLE DISPLAY
        if rows:
            id_col_name = "ரசீது எண்" if "வரவு" in view_type else "செலவு எண்"
            display_data = []

            for idx, r in enumerate(rows, start=1):
                p_val = r[5] if len(r) > 5 and r[5] else "-"
                p_method = (
                    r[6]
                    if len(r) > 6 and r[6]
                    else ("-" if "செலவு" in view_type else "Cash (பணம்)")
                )

                if "வரவு" in view_type:
                    display_data.append(
                        [idx, r[0], r[1], r[2], r[3], p_method, r[4], p_val]
                    )
                else:
                    display_data.append(
                        [idx, r[0], r[1], r[2], r[3], "-", r[4], p_val]
                    )

            columns = [
                "வ.எண் (S.No)",
                id_col_name,
                "தேதி",
                "வகை",
                "பெயர் / விவரம்",
                "செலுத்திய முறை",
                "தொகை (₹)",
                "கைபேசி / குறிப்பு",
            ]
            df = pd.DataFrame(display_data, columns=columns)
            st.dataframe(df, use_container_width=True)

            excel_data = generate_excel_report(
                rows, view_type, cat_filter, filtered_tot
            )
            st.download_button(
                label="📊 Excel அறிக்கையாகப் பதிவிறக்கு",
                data=excel_data,
                file_name=f"Kovil_Report_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        else:
            st.info("தகவல்கள் எதுவும் இல்லை (No records found).")

    # TAB 4: EDIT ENTRY
    with tab4:
        st.header("✏️ பதிவை மீண்டும் திருத்துதல் (Edit Entry)")

        edit_type = st.radio(
            "எதை திருத்த வேண்டும்?:",
            ["📥 வரவு (Receipt)", "📤 செலவு (Expense)"],
            horizontal=True,
        )

        edit_id = st.number_input(
            "திருத்த வேண்டிய எண் (Receipt No / Expense ID):",
            min_value=1,
            step=1,
        )

        if st.button("🔍 தகவலைக் கொண்டுவா"):
            with get_db_connection() as conn:
                cursor = conn.cursor()
                if "வரவு" in edit_type:
                    cursor.execute(
                        "SELECT receipt_no, date, category, name, city, amount,"
                        " phone, payment_method FROM receipts WHERE receipt_no"
                        " = ?",
                        (edit_id,),
                    )
                else:
                    cursor.execute(
                        "SELECT expense_id, date, category, title, amount,"
                        " remarks FROM expenses WHERE expense_id = ?",
                        (edit_id,),
                    )
                st.session_state["edit_data"] = cursor.fetchone()

        if "edit_data" in st.session_state and st.session_state["edit_data"]:
            data = st.session_state["edit_data"]
            st.info(f"எண் {data[0]}-ன் பழைய தகவல்கள் மாற்றத்திற்குத் தயார்.")

            with st.form("edit_form"):
                try:
                    p_date = datetime.strptime(data[1], "%d-%m-%Y").date()
                except Exception:
                    p_date = date.today()

                e_date = st.date_input(
                    "புதிய தேதி:",
                    value=p_date,
                    min_value=date(1900, 1, 1),
                    max_value=date(2100, 12, 31),
                )

                if "வரவு" in edit_type:
                    cat_list = [
                        "சிறப்பு வைப்பு நிதியாளர்கள்",
                        "சிறப்பு காணிக்கையாளர்கள்",
                        "காணிக்கையாளர்கள் வரவுகள்",
                        "காணிக்கையாளர்கள்",
                    ]
                    e_cat = st.selectbox(
                        "வரவு வகை:",
                        cat_list,
                        index=cat_list.index(data[2])
                        if data[2] in cat_list
                        else 0,
                    )
                    e_name = st.text_input("பெயர்:", value=data[3])
                    e_city = st.text_input("ஊர்:", value=data[4])
                    e_amt = st.number_input(
                        "தொகை (₹):", value=float(data[5]), step=50.0
                    )
                    e_phone = st.text_input("கைபேசி:", value=data[6] or "")

                    pay_methods = [
                        "Cash (பணம்)",
                        "GPay / PhonePe (UPI)",
                        "Bank Transfer (வங்கி மாற்றம்)",
                    ]
                    cur_method = (
                        data[7] if len(data) > 7 and data[7] else "Cash (பணம்)"
                    )
                    e_method = st.selectbox(
                        "செலுத்திய முறை:",
                        pay_methods,
                        index=pay_methods.index(cur_method)
                        if cur_method in pay_methods
                        else 0,
                    )
                else:
                    cat_list = [
                        "மின்சாரக் கட்டணம் (EB Bill)",
                        "பூஜைப் பொருட்கள்",
                        "அன்னதானம்",
                        "வேலை ஆள் கூலி",
                        "பராமரிப்புச் செலவு",
                        "இதர செலவுகள்",
                    ]
                    e_cat = st.selectbox(
                        "செலவு வகை:",
                        cat_list,
                        index=cat_list.index(data[2])
                        if data[2] in cat_list
                        else 0,
                    )
                    e_title = st.text_input("விவரம்:", value=data[3])
                    e_amt = st.number_input(
                        "தொகை (₹):", value=float(data[4]), step=50.0
                    )
                    e_remarks = st.text_input("குறிப்பு:", value=data[5] or "")

                update_btn = st.form_submit_button(
                    "🔄 மாற்றங்களை சேமி (Update)", type="primary"
                )

                if update_btn:
                    formatted_u_date = e_date.strftime("%d-%m-%Y")
                    with get_db_connection() as conn:
                        cursor = conn.cursor()
                        if "வரவு" in edit_type:
                            cursor.execute(
                                "UPDATE receipts SET date=?, category=?,"
                                " name=?, city=?, amount=?, phone=?,"
                                " payment_method=? WHERE receipt_no=?",
                                (
                                    formatted_u_date,
                                    e_cat,
                                    e_name,
                                    e_city,
                                    e_amt,
                                    e_phone,
                                    e_method,
                                    data[0],
                                ),
                            )
                        else:
                            cursor.execute(
                                "UPDATE expenses SET date=?, category=?,"
                                " title=?, amount=?, remarks=? WHERE"
                                " expense_id=?",
                                (
                                    formatted_u_date,
                                    e_cat,
                                    e_title,
                                    e_amt,
                                    e_remarks,
                                    data[0],
                                ),
                            )
                        conn.commit()

                    st.success("✅ தகவல்கள் வெற்றிகரமாக புதுப்பிக்கப்பட்டன!")
                    del st.session_state["edit_data"]
                    st.rerun()
